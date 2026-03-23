import re
import time
from dataclasses import dataclass
from datetime import datetime
from typing import List, Set, Tuple, Dict
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from serpapi import GoogleSearch


SERP_API_KEY = "7c1b43ef71cf518e6e0d52d814b867277b264f16fd27ca59abb38ac4acfcaa0e"


@dataclass
class ScrapeConfig:
    max_workers: int = 5
    request_timeout: int = 12
    delay_between_pages_sec: float = 1.0
    delay_between_site_pages_sec: float = 0.4
    results_per_page: int = 10
    max_results_per_query: int = 100
    max_sites_to_scrape_per_query: int = 100
    only_save_with_both_email_and_phone: bool = True
    pages_to_check: Tuple[str, ...] = (
        "",
        "/contact",
        "/contact-us",
        "/about",
        "/about-us",
        "/get-in-touch",
    )
    exclude_domains: Tuple[str, ...] = (
        "facebook.com",
        "instagram.com",
        "yelp.com",
        "tripadvisor.com",
        "linkedin.com",
        "opentable.com",
        "ubereats.com",
        "doordash.com",
        "youtube.com",
        "tiktok.com",
        "snapchat.com"
    )


class LeadGenerationScraper:
    EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")

    # AU landline/mobile-like formats
    AU_PHONE_REGEX = re.compile(
        r"(?:\+61\s?[2-9]\s?\d{4}\s?\d{4}|0[2-9]\s?\d{4}\s?\d{4})"
    )

    def __init__(self, api_key: str, config: ScrapeConfig = ScrapeConfig()):
        self.api_key = api_key
        self.cfg = config
        self.results: List[Dict] = []

        self.headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
        }

        self.session = requests.Session()
        self.executor = ThreadPoolExecutor(max_workers=self.cfg.max_workers)

    def _normalize_email(self, email: str) -> str:
        if not email:
            return ""
        email = email.strip().strip(".,;:!?)[]{}<>\"'").lower()
        return email

    def _normalize_phone(self, phone: str) -> str:
        if not phone:
            return ""
        phone = re.sub(r"\s+", " ", phone).strip()
        return phone

    def _is_excluded(self, url: str) -> bool:
        u = url.lower()
        return any(domain in u for domain in self.cfg.exclude_domains)

    def _canonicalize_url(self, url: str) -> str:
        try:
            p = urlparse(url)
            scheme = p.scheme or "https"
            netloc = p.netloc
            path = p.path or "/"
            return f"{scheme}://{netloc}{path}".rstrip("/")
        except Exception:
            return url

    def _is_valid_email(self, email: str) -> bool:
        if not email:
            return False
        if not self.EMAIL_REGEX.fullmatch(email):
            return False

        bad_parts = [
            ".png", ".jpg", ".jpeg", ".webp", ".svg", ".gif",
            ".css", ".js", ".ico", "@example.com"
        ]
        email_lower = email.lower()
        if any(part in email_lower for part in bad_parts):
            return False

        return True

    def _is_valid_phone(self, phone: str) -> bool:
        if not phone:
            return False

        phone = self._normalize_phone(phone)
        digits = re.sub(r"\D", "", phone)

        # reject obvious dates / junk
        if len(digits) < 8 or len(digits) > 15:
            return False

        # prefer AU formats strongly
        if self.AU_PHONE_REGEX.search(phone):
            return True

        # also allow normalized tel values that look realistic
        if digits.startswith("61") and len(digits) in (10, 11):
            return True
        if digits.startswith("0") and len(digits) == 10:
            return True

        return False

    def get_websites(self, query: str) -> List[str]:
        websites: List[str] = []
        seen: Set[str] = set()

        target = min(
            self.cfg.max_results_per_query,
            self.cfg.max_sites_to_scrape_per_query
        )
        step = self.cfg.results_per_page

        for start in range(0, target, step):
            params = {
                "engine": "google",
                "q": query,
                "api_key": self.api_key,
                "num": step,
                "start": start,
            }

            data = GoogleSearch(params).get_dict()
            organic = data.get("organic_results", [])
            if not organic:
                break

            for result in organic:
                link = result.get("link")
                if not link:
                    continue
                if self._is_excluded(link):
                    continue

                clean_url = self._canonicalize_url(link)
                if clean_url not in seen:
                    seen.add(clean_url)
                    websites.append(clean_url)

                if len(websites) >= self.cfg.max_sites_to_scrape_per_query:
                    break

            if len(websites) >= self.cfg.max_sites_to_scrape_per_query:
                break

            time.sleep(self.cfg.delay_between_pages_sec)

        return websites

    def extract_contacts_from_html(self, html: str) -> Tuple[List[str], List[str]]:
        soup = BeautifulSoup(html, "html.parser")

        emails: Set[str] = set()
        phones: Set[str] = set()

        # mailto / tel first
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()

            if href.lower().startswith("mailto:"):
                email = href.split(":", 1)[1].split("?", 1)[0]
                email = self._normalize_email(email)
                if self._is_valid_email(email):
                    emails.add(email)

            if href.lower().startswith("tel:"):
                phone = href.split(":", 1)[1].split("?", 1)[0]
                phone = self._normalize_phone(phone)
                if self._is_valid_phone(phone):
                    phones.add(phone)

        text = soup.get_text(" ", strip=True)

        for email in self.EMAIL_REGEX.findall(text):
            email = self._normalize_email(email)
            if self._is_valid_email(email):
                emails.add(email)

        for phone in self.AU_PHONE_REGEX.findall(text):
            phone = self._normalize_phone(phone)
            if self._is_valid_phone(phone):
                phones.add(phone)

        clean_emails = sorted({e for e in emails if self._is_valid_email(e)})
        clean_phones = sorted({p for p in phones if self._is_valid_phone(p)})

        return clean_emails, clean_phones

    def scrape_website(self, base_url: str) -> Tuple[List[str], List[str]]:
        all_emails: Set[str] = set()
        all_phones: Set[str] = set()

        for path in self.cfg.pages_to_check:
            try:
                full_url = urljoin(base_url + "/", path.lstrip("/"))
                resp = self.session.get(
                    full_url,
                    headers=self.headers,
                    timeout=self.cfg.request_timeout
                )

                if resp.status_code != 200 or not resp.text:
                    continue

                emails, phones = self.extract_contacts_from_html(resp.text)

                all_emails.update(emails)
                all_phones.update(phones)

                time.sleep(self.cfg.delay_between_site_pages_sec)

            except Exception:
                continue

        clean_emails = sorted({e for e in all_emails if self._is_valid_email(e)})
        clean_phones = sorted({p for p in all_phones if self._is_valid_phone(p)})

        return clean_emails, clean_phones

    def process_website(self, url: str, query: str) -> Dict:
        print(f"Scraping: {url}")
        emails, phones = self.scrape_website(url)

        return {
            "search_query": query,
            "website": url,
            "emails": emails,
            "phones": phones,
            "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def scrape(self, query: str):
        websites = self.get_websites(query)
        print(f"\nFound {len(websites)} websites for '{query}'\n")

        futures = [
            self.executor.submit(self.process_website, site, query)
            for site in websites
        ]

        for future in as_completed(futures):
            item = future.result()

            item["emails"] = [e for e in item["emails"] if self._is_valid_email(e)]
            item["phones"] = [p for p in item["phones"] if self._is_valid_phone(p)]

            if self.cfg.only_save_with_both_email_and_phone:
                if not item["emails"] or not item["phones"]:
                    continue

            self.results.append(item)

    def _deduplicate_results(self) -> List[Dict]:
        best_by_site: Dict[str, Dict] = {}

        for row in self.results:
            site = row["website"]
            score = len(row["emails"]) + len(row["phones"])

            if site not in best_by_site:
                best_by_site[site] = row
            else:
                old = best_by_site[site]
                old_score = len(old["emails"]) + len(old["phones"])
                if score > old_score:
                    best_by_site[site] = row

        return list(best_by_site.values())

    def save_excel(self, filename: str = "australia_business_leads.xlsx"):
        final = self._deduplicate_results()

        # keep only rows with both email and phone
        final = [
            row for row in final
            if row["emails"] and row["phones"]
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers = [
            "Search Query",
            "Website",
            "Emails",
            "Phones",
            "Email Count",
            "Phone Count",
            "Scraped At",
        ]
        ws.append(headers)

        for row in final:
            ws.append([
                row["search_query"],
                row["website"],
                ", ".join(row["emails"]),
                ", ".join(row["phones"]),
                len(row["emails"]),
                len(row["phones"]),
                row["scraped_at"],
            ])

        # styling
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # widths
        widths = {
            1: 28,
            2: 50,
            3: 45,
            4: 25,
            5: 12,
            6: 12,
            7: 22,
        }
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # wrap long text
        for row in ws.iter_rows(min_row=2):
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
            row[3].alignment = Alignment(wrap_text=True, vertical="top")
            row[1].alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"
        wb.save(filename)
        print(f"\nSaved {len(final)} cleaned leads to {filename}")

    def close(self):
        try:
            self.executor.shutdown(wait=True)
        except Exception:
            pass
        try:
            self.session.close()
        except Exception:
            pass


if __name__ == "__main__":
    scraper = LeadGenerationScraper(
        SERP_API_KEY,
        ScrapeConfig(
            max_workers=5,
            max_results_per_query=100,
            max_sites_to_scrape_per_query=100,
            only_save_with_both_email_and_phone=True,
        )
    )

    search_queries = [
        "Childcare in Bondi Junction",
        "Childcare in Bonnyrigg",
        "Childcare in Brookvale",
        "Childcare in Burwood",
        "Childcare in Cabramatta",
        "Childcare in Camden",
        "Childcare in Campbelltown",
        "Childcare in Caringbah",
        "Childcare in Castle Hill",
        "Childcare in Casula"
    ]

    try:
        for query in search_queries:
            scraper.scrape(query)

        scraper.save_excel("australia_business_leads.xlsx")
    finally:
        scraper.close()